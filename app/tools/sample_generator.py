"""
Sample BOQ & CAD File Generator for Vertex Construction & PCCC Testing
Generates realistic commercial Fire Protection & MEP equipment BOQs.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import ezdxf


def create_sample_excel_boq(file_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ Thiet Bi PCCC & Co Dien"

    headers = ["STT", "Tên thiết bị / Vật tư", "Quy cách / Thông số kỹ thuật", "Độ dày / Model", "ĐVT", "Số lượng", "Ghi chú"]
    
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "BẢNG TỔNG HỢP KHỐI LƯỢNG THIẾT BỊ PCCC & CƠ ĐIỆN - DỰ ÁN TÒA NHÀ PHỨC HỢP"
    ws["A1"].font = Font(name="Arial", size=12, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Header
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1B2234", end_color="1B2234", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data: Comprehensive PCCC & HVAC Equipment
    sample_data = [
        (1, "Bình chữa cháy bột ABC 4kg có kiểm định", "Model: MFZL4, dán tem BCA", 0, "bình", 30.0, "Tem kiểm định PCCC"),
        (2, "Bình chữa cháy bột ABC 8kg có kiểm định", "Model: MFZL8, dán tem BCA", 0, "bình", 15.0, "Tem kiểm định PCCC"),
        (3, "Bình chữa cháy khí CO2 3kg phòng điện", "Model: MT3, khí CO2 tinh khiết", 0, "bình", 12.0, "Phòng server & tủ điện"),
        (4, "Bình chữa cháy xe đẩy bột ABC 35kg", "Model: MFTZ35 có bánh xe", 0, "bình", 2.0, "Đặt tại tầng hầm kho"),
        (5, "Đầu báo khói quang điện địa chỉ 24V", "Kèm đế tiêu chuẩn, LED hiển thị", 0, "bộ", 48.0, "Đạt chuẩn TCVN 5738"),
        (6, "Đầu báo nhiệt gia tăng và cố định 57°C", "Kèm đế gắn trần", 0, "bộ", 18.0, "Khu vực bếp và hầm xe"),
        (7, "Đèn chỉ dẫn thoát hiểm Exit LED 2 mặt", "Pin dự phòng 120 phút tự sạc", 0, "bộ", 26.0, "Lối thoát hiểm cầu thang"),
        (8, "Đèn chiếu sáng sự cố khẩn cấp 2 mắt LED", "Emergency light tự động bật", 0, "bộ", 24.0, "Hành lang & sảnh thang"),
        (9, "Đầu phun chữa cháy tự động Sprinkler 68°C", "DN15 K=5.6 hướng xuống", 0, "cái", 120.0, "Bóng thủy tinh đỏ"),
        (10, "Trụ chữa cháy 3 cửa ngoài trời DN100", "Gang xám (2xD65 + 1xD100)", 0, "bộ", 2.0, "Áp lực làm việc 1.6 MPa"),
        (11, "Cuộn vòi chữa cháy D50 dài 20m kèm khớp nối", "Áp lực thử 2.5 MPa", 0, "cuộn", 16.0, "Hộp chữa cháy vách tường"),
        (12, "Van ngăn cháy cầu chì tự đóng 70°C (FD)", "500x300mm bích TDC", 1.2, "cái", 8.0, "Tự đóng khi đạt 70 độ C"),
        (13, "Ống gió chống cháy bọc tấm cách nhiệt EI 60", "500x300mm L=1200", 0.75, "m2", 65.0, "Hệ thống hút khói PCCC"),
        (14, "Ống gió thẳng vuông tôn mạ kẽm", "500x300 L=1200", 0.75, "m2", 85.0, "Bích TDC kèm nẹp"),
        (15, "Cửa gió khuếch tán 4 hướng kèm hộp gió", "600x600 nhôm sơn tĩnh điện", 1.0, "bộ", 20.0, "Màu trắng RAL 9010")
    ]

    for row_idx, row_data in enumerate(sample_data, start=4):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)
    wb.close()
    return file_path


def create_sample_cad_dxf(file_path: str):
    doc = ezdxf.new("R2010")
    
    # Define engineering layers
    doc.layers.add("PCCC_SPRINKLER_PENDENT", color=1)
    doc.layers.add("PCCC_PIPE_DN100", color=3)
    doc.layers.add("HVAC_DUCT_SUPPLY", color=4)
    doc.layers.add("PCCC_ALARM_SMOKE", color=5)
    doc.layers.add("GHI_CHU", color=7)
    
    msp = doc.modelspace()

    # Add geometric pipe lines (e.g. DN100 100m total = 100,000mm)
    msp.add_line((0, 0), (50000, 0), dxfattribs={"layer": "PCCC_PIPE_DN100"})
    msp.add_line((50000, 0), (50000, 50000), dxfattribs={"layer": "PCCC_PIPE_DN100"})
    
    # Add sprinkler circles
    for i in range(10):
        msp.add_circle((i * 5000, 10000), radius=15, dxfattribs={"layer": "PCCC_SPRINKLER_PENDENT"})

    # Add text entities representing PCCC and MEP equipment in CAD drawing
    texts = [
        "BÌNH CHỮA CHÁY MFZL4 - 30 BÌNH",
        "BÌNH CO2 MT3 - 12 BÌNH",
        "ĐẦU BÁO KHÓI QUANG 24V - 48 BỘ",
        "ĐẦU BÁO NHIỆT GIA TĂNG - 18 BỘ",
        "ĐÈN EXIT LED 2 MẶT - 26 BỘ",
        "ĐÈN SỰ CỐ EMERGENCY - 24 BỘ",
        "SPRINKLER DN15 68C - 120 CÁI",
        "TRỤ CỨU HỎA 3 CỬA DN100 - 2 BỘ",
        "VÒI CHỮA CHÁY D50 20M - 16 CUỘN",
        "VAN NGĂN CHÁY FD 500x300 - 8 CÁI",
        "ỐNG GIÓ CHỐNG CHÁY EI60 500x300 - 65 M2"
    ]

    for idx, t in enumerate(texts):
        msp.add_text(t, dxfattribs={"height": 2.5, "insert": (10, idx * 15), "layer": "GHI_CHU"})

    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    doc.saveas(file_path)
    return file_path


if __name__ == "__main__":
    p1 = "storage/samples/BOQ_Mau_Ong_Gio_Vertex.xlsx"
    p2 = "storage/samples/Ban_Ve_CAD_Ong_Gio.dxf"
    create_sample_excel_boq(p1)
    create_sample_cad_dxf(p2)
    print("Created PCCC sample files successfully.")
