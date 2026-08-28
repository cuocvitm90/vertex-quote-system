"""
Master Template Generator for Vertex Construction & PCCC
Generates the official company Master Template Excel file containing standard items,
cost breakdown structure, and pricing coefficient framework (% waste, transport, labor, margin).
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.config import settings

def create_master_template_excel(output_path: str = "storage/templates/Master_Template_Vertex.xlsx") -> str:
    """Generates a branded Master Template Excel spreadsheet"""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: Master Catalog
    ws1 = wb.active
    ws1.title = "Danh Mục Vật Tư Chuẩn"
    
    # Header styling
    navy_fill = PatternFill(start_color="1B2234", end_color="1B2234", fill_type="solid")
    orange_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    subtle_fill = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
    
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_title = Font(name="Arial", size=14, bold=True, color="1B2234")
    font_bold = Font(name="Arial", size=10, bold=True)
    font_regular = Font(name="Arial", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Title
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "CÔNG TY CỔ PHẦN VERTEX CONSTRUCTION & PCCC - BẢNG DANH MỤC VẬT TƯ CHUẨN (MASTER CATALOG)"
    ws1["A1"].font = font_title
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 35
    
    headers = ["Mã Vật Tư", "Tên Vật Tư / Thiết Bị Chuẩn", "Phân Loại", "Quy Cách / Kích Thước", "ĐVT", "Đơn Giá Gốc (VNĐ)", "Ghi Chú"]
    ws1.append([])
    ws1.append(headers)
    ws1.row_dimensions[3].height = 26
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col_idx)
        cell.fill = navy_fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    sample_items = [
        ["PCCC-BCC-ABC4", "Bình chữa cháy bột ABC 4kg (MFZL4)", "Bình chữa cháy", "4kg có tem BCA", "bình", 280000, "Chuẩn kiểm định"],
        ["PCCC-BCC-ABC8", "Bình chữa cháy bột ABC 8kg (MFZL8)", "Bình chữa cháy", "8kg có tem BCA", "bình", 385000, "Chuẩn kiểm định"],
        ["PCCC-BAOKHOI-QUANG", "Đầu báo khói quang điện địa chỉ 24V", "Báo cháy", "Kèm đế tiêu chuẩn", "bộ", 320000, "Hệ thống báo cháy"],
        ["PCCC-SPK-68C", "Đầu phun chữa cháy tự động Sprinkler 68°C", "Chữa cháy nước", "DN15 K=5.6 hướng xuống", "cái", 65000, "Đạt chuẩn UL/FM"],
        ["PCCC-EXIT-LED", "Đèn Exit LED thoát hiểm 2 mặt dạ quang", "Đèn thoát hiểm", "Pin dự phòng 120 phút", "bộ", 260000, "Tự động sạc"],
        ["HVAC-ONG-VUONG-075", "Ống gió vuông tôn mạ kẽm d=0.75mm", "Ống gió", "Tôn mạ kẽm Hoa Sen bích TDC", "m2", 245000, "Tiêu chuẩn SMACNA"],
        ["HVAC-CUT-90-075", "Cút 90 độ ống gió vuông d=0.75mm", "Phụ kiện ống gió", "Bích TDC kèm nẹp", "m2", 285000, "Gia công CNC"],
        ["HVAC-VAN-FD-500", "Van ngăn cháy cầu chì nhiệt FD 70°C", "Van gió PCCC", "500x300mm thân tôn tráng kẽm", "cái", 580000, "Cầu chì tự ngắt"]
    ]

    for row_idx, item in enumerate(sample_items, 4):
        ws1.append(item)
        ws1.row_dimensions[row_idx].height = 20
        for col_idx in range(1, 8):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            if col_idx in [1, 5]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 6:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column widths
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 42
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 28
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 20
    ws1.column_dimensions['G'].width = 25

    # Sheet 2: Pricing Coefficients Framework
    ws2 = wb.create_sheet(title="Khung Định Mức & Hệ Số")
    
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "KHUNG HỆ SỐ ĐỊNH MỨC CHI PHÍ VẬT TƯ (PRICING COEFFICIENTS)"
    ws2["A1"].font = font_title
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 35
    
    c_headers = ["STT", "Hạng Mục Định Mức Hệ Số", "Tỷ Lệ Mặc Định (%)", "Diễn Giải Kỹ Thuật"]
    ws2.append([])
    ws2.append(c_headers)
    ws2.row_dimensions[3].height = 26
    
    for col_idx, h in enumerate(c_headers, 1):
        cell = ws2.cell(row=3, column=col_idx)
        cell.fill = orange_fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    coefficients_data = [
        [1, "Tỷ lệ hao hụt vật tư & cắt gọt tại công trình (Waste Factor)", 0.05, "Bù hao hụt tôn, bu lông, roong, que hàn khi thi công"],
        [2, "Tỷ lệ vận chuyển & bốc dỡ (Transport / Logistics)", 0.03, "Vận chuyển từ nhà máy Vertex đến chân công trình"],
        [3, "Tỷ lệ nhân công chế tạo & phụ kiện lắp ghép (Labor / Accessories)", 0.15, "Nhân công gò, ghép bích, hàn, tán đinh, sơn chống gỉ"],
        [4, "Biên độ lợi nhuận gộp định mức (Profit Margin)", 0.12, "Biên độ lợi nhuận chuẩn của Vertex"],
        [5, "TỔNG HỆ SỐ NHÂN ÁP DỤNG VÀO GIÁ THÔ (MARKUP MULTIPLIER)", 0.35, "Công thức: Giá Bán = Giá Thô x (1 + 0.05 + 0.03 + 0.15 + 0.12) = Giá Thô x 1.35"]
    ]

    for row_idx, row_data in enumerate(coefficients_data, 4):
        ws2.append(row_data)
        ws2.row_dimensions[row_idx].height = 24
        is_total = (row_idx == 8)
        for col_idx in range(1, 5):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.font = font_bold if is_total else font_regular
            cell.border = thin_border
            if is_total:
                cell.fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 3:
                cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 58
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 65

    wb.save(output_path)
    return output_path
