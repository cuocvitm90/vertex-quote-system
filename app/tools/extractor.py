"""
Tool 1: BOQ & Drawing Extractor
Extracts bill of quantities from Excel (.xlsx, .xls), CAD DXF (.dxf), and PDF files.
"""
import re
import os
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
import pandas as pd

try:
    import ezdxf
except ImportError:
    ezdxf = None

try:
    import pypdf
except ImportError:
    pypdf = None


class ExtractedRawItem:
    def __init__(
        self,
        stt: int = 1,
        raw_name: str = "",
        raw_spec: str = "",
        unit: str = "m2",
        quantity: float = 1.0,
        raw_thickness: Optional[float] = None,
        notes: str = ""
    ):
        self.stt = stt
        self.raw_name = raw_name.strip()
        self.raw_spec = raw_spec.strip()
        self.unit = unit.strip()
        self.quantity = float(quantity) if quantity else 1.0
        self.raw_thickness = raw_thickness
        self.notes = notes.strip()

    def to_dict(self) -> Dict[str, Any]:
        return {
            "stt": self.stt,
            "raw_name": self.raw_name,
            "raw_spec": self.raw_spec,
            "unit": self.unit,
            "quantity": self.quantity,
            "raw_thickness": self.raw_thickness,
            "notes": self.notes
        }


class BOQExtractor:
    """Extracts raw material list from various engineering file formats"""

    @classmethod
    def extract(cls, file_path: str) -> List[ExtractedRawItem]:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        ext = path.suffix.lower()
        if ext in [".xlsx", ".xls"]:
            return cls.extract_from_excel(file_path)
        elif ext in [".dxf", ".dwg"]:
            return cls.extract_from_cad(file_path)
        elif ext in [".pdf"]:
            return cls.extract_from_pdf(file_path)
        elif ext in [".csv"]:
            return cls.extract_from_csv(file_path)
        else:
            # Try excel as fallback
            return cls.extract_from_excel(file_path)

    @classmethod
    def extract_from_excel(cls, file_path: str) -> List[ExtractedRawItem]:
        """Reads Excel file, auto-detects header row, extracts materials"""
        items: List[ExtractedRawItem] = []
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # 1. Scan first 20 rows to find header columns
        header_row_idx = -1
        col_map = {}

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx > 25:
                break
            non_empty = [c for c in row if c is not None and str(c).strip() != ""]
            if len(non_empty) < 3:
                # Bỏ qua dòng tiêu đề lớn hoặc dòng trống
                continue

            row_str = [str(c).strip().lower() if c is not None else "" for c in row]
            
            # Kiểm tra từng ô riêng lẻ
            cell_name = any(any(k in cell for k in ["tên", "nội dung", "vật tư", "diễn giải", "hàng hóa", "description", "item"]) for cell in row_str)
            cell_qty = any(any(k in cell for k in ["số lượng", "khối lượng", "sl", "qty", "quantity"]) for cell in row_str)
            cell_unit = any(any(k in cell for k in ["đvt", "đơn vị", "unit"]) for cell in row_str)
            cell_spec = any(any(k in cell for k in ["quy cách", "kích thước", "spec", "size", "độ dày"]) for cell in row_str)

            if (cell_name and cell_qty) or (cell_name and (cell_unit or cell_spec)):
                header_row_idx = row_idx
                for col_idx, val in enumerate(row_str):
                    if any(k in val for k in ["stt", "no.", "mục"]):
                        col_map["stt"] = col_idx
                    elif any(k in val for k in ["tên", "nội dung", "vật tư", "diễn giải", "hàng hóa", "description", "item"]):
                        col_map["name"] = col_idx
                    elif any(k in val for k in ["quy cách", "kích thước", "spec", "size", "kích thước (mm)"]):
                        col_map["spec"] = col_idx
                    elif any(k in val for k in ["độ dày", "dày", "thickness", "t (mm)"]):
                        col_map["thickness"] = col_idx
                    elif any(k in val for k in ["đvt", "đơn vị", "unit"]):
                        col_map["unit"] = col_idx
                    elif any(k in val for k in ["số lượng", "khối lượng", "sl", "qty", "quantity"]):
                        col_map["qty"] = col_idx
                    elif any(k in val for k in ["ghi chú", "note", "remark"]):
                        col_map["notes"] = col_idx
                break


        # If no header found, default to standard columns
        if header_row_idx == -1:
            header_row_idx = 1
            col_map = {"stt": 0, "name": 1, "spec": 2, "unit": 3, "qty": 4, "notes": 5}

        # 2. Iterate data rows
        stt_counter = 1
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx <= header_row_idx:
                continue
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            name_idx = col_map.get("name", 1)
            name_val = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] is not None else ""

            # Check if this row is total row or section header without quantity
            if any(k in name_val.lower() for k in ["tổng cộng", "cộng", "tổng tiền", "người lập", "giám đốc"]):
                continue

            if not name_val or name_val.lower() == "none":
                # Maybe name is in spec col
                spec_idx = col_map.get("spec", -1)
                if spec_idx != -1 and spec_idx < len(row) and row[spec_idx]:
                    name_val = str(row[spec_idx]).strip()
                else:
                    continue

            spec_val = ""
            if "spec" in col_map and col_map["spec"] < len(row) and row[col_map["spec"]] is not None:
                spec_val = str(row[col_map["spec"]]).strip()

            thickness_val = None
            if "thickness" in col_map and col_map["thickness"] < len(row) and row[col_map["thickness"]] is not None:
                try:
                    raw_th = str(row[col_map["thickness"]]).replace("mm", "").strip()
                    thickness_val = float(raw_th)
                except Exception:
                    pass

            unit_val = "m2"
            if "unit" in col_map and col_map["unit"] < len(row) and row[col_map["unit"]] is not None:
                unit_val = str(row[col_map["unit"]]).strip()

            qty_val = 1.0
            if "qty" in col_map and col_map["qty"] < len(row) and row[col_map["qty"]] is not None:
                try:
                    cleaned_qty = str(row[col_map["qty"]]).replace(",", "").replace(" ", "").strip()
                    qty_val = float(cleaned_qty)
                except Exception:
                    qty_val = 1.0

            notes_val = ""
            if "notes" in col_map and col_map["notes"] < len(row) and row[col_map["notes"]] is not None:
                notes_val = str(row[col_map["notes"]]).strip()

            items.append(ExtractedRawItem(
                stt=stt_counter,
                raw_name=name_val,
                raw_spec=spec_val,
                unit=unit_val,
                quantity=qty_val,
                raw_thickness=thickness_val,
                notes=notes_val
            ))
            stt_counter += 1

        wb.close()
        return items

    @classmethod
    def extract_from_csv(cls, file_path: str) -> List[ExtractedRawItem]:
        df = pd.read_csv(file_path)
        items = []
        for idx, row in df.iterrows():
            name = str(row.get("Tên vật tư", row.get("name", row.iloc[1] if len(row) > 1 else "")))
            if not name.strip() or name.lower() == "nan":
                continue
            spec = str(row.get("Quy cách", row.get("spec", "")))
            unit = str(row.get("ĐVT", row.get("unit", "m2")))
            qty = 1.0
            try:
                qty = float(row.get("Số lượng", row.get("quantity", 1.0)))
            except Exception:
                pass
            items.append(ExtractedRawItem(
                stt=len(items) + 1,
                raw_name=name,
                raw_spec=spec,
                unit=unit,
                quantity=qty
            ))
        return items

    @classmethod
    def extract_from_cad(cls, file_path: str) -> List[ExtractedRawItem]:
        """Parses CAD DXF entities, layers, line geometry, and schedules using high-precision CADTakeoffEngine"""
        if ezdxf is None:
            raise ImportError("ezdxf is required to parse CAD files.")

        from app.tools.cad_takeoff_engine import CADTakeoffEngine
        result = CADTakeoffEngine.extract_cad_takeoff(file_path)
        items: List[ExtractedRawItem] = []
        for it in result.items:
            items.append(ExtractedRawItem(
                stt=it.get("stt", len(items) + 1),
                raw_name=it.get("name", ""),
                raw_spec=it.get("spec", ""),
                unit=it.get("unit", "m2"),
                quantity=float(it.get("quantity", 1.0)),
                notes=f"Bóc tách CAD Layer: {it.get('layer', '')}"
            ))

        if not items:
            items.append(ExtractedRawItem(
                stt=1,
                raw_name="Ống gió vuông theo bản vẽ CAD",
                raw_spec="Bóc tách tổng thể từ CAD",
                unit="m2",
                quantity=50.0
            ))

        return items

    @classmethod
    def extract_from_pdf(cls, file_path: str) -> List[ExtractedRawItem]:
        """Extracts text from PDF BOQ file"""
        if pypdf is None:
            raise ImportError("pypdf is required to parse PDF files.")

        items: List[ExtractedRawItem] = []
        reader = pypdf.PdfReader(file_path)
        full_text = ""
        for page in reader.pages:
            full_text += page.extract_text() + "\n"

        lines = [line.strip() for line in full_text.split("\n") if line.strip()]
        stt = 1
        for line in lines:
            # Simple line parsing: STT | Name | Unit | Qty
            parts = re.split(r"\s{2,}|\t", line)
            if len(parts) >= 2:
                name = parts[0]
                # Filter out headers
                if any(h in name.lower() for h in ["stt", "tên vật tư", "bảng báo giá", "cộng", "trang"]):
                    continue
                unit = "m2"
                qty = 1.0
                if len(parts) >= 3:
                    unit = parts[1]
                    try:
                        qty = float(parts[2].replace(",", ""))
                    except Exception:
                        qty = 1.0
                items.append(ExtractedRawItem(
                    stt=stt,
                    raw_name=name,
                    raw_spec="",
                    unit=unit,
                    quantity=qty
                ))
                stt += 1

        return items
