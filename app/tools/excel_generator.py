"""
Tool 4: Professional Excel Quotation Generator for Vertex Construction & PCCC
Uses openpyxl to generate high-quality branded quotation workbooks supporting Multi-language (vi, en, zh, ko).
"""
from pathlib import Path
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.database.models import Quote
from app.config import settings
from app.services.i18n import t


class VertexExcelGenerator:
    """Tạo file Excel Báo giá chuẩn form mẫu chuyên nghiệp Vertex Construction & PCCC (Đa ngôn ngữ)"""

    @classmethod
    def generate_quote_excel(cls, quote: Quote, output_path: Optional[str] = None) -> str:
        return cls.generate(quote, output_path)

    @classmethod
    def generate(cls, quote: Quote, output_path: Optional[str] = None) -> str:
        lang = getattr(quote, "language", "vi") or "vi"

        if not output_path:
            filename = f"Bao_Gia_{quote.quote_code}_{quote.customer_name.replace(' ', '_')}_{lang}.xlsx"
            clean_name = "".join(c for c in filename if c.isalnum() or c in "._-")
            output_path = str(Path(settings.QUOTES_DIR) / clean_name)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Vertex Quote ({lang.upper()})"
        ws.views.sheetView[0].showGridLines = True

        FONT_FAMILY = "Segoe UI"
        font_header_company = Font(name=FONT_FAMILY, size=13, bold=True, color="1B2234")
        font_sub_company = Font(name=FONT_FAMILY, size=9, italic=True, color="64748B")
        font_title = Font(name=FONT_FAMILY, size=15, bold=True, color="1B2234")
        font_bold = Font(name=FONT_FAMILY, size=10, bold=True, color="1E293B")
        font_regular = Font(name=FONT_FAMILY, size=10, color="334155")
        font_italic = Font(name=FONT_FAMILY, size=9, italic=True, color="64748B")
        font_table_header = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        font_total_bold = Font(name=FONT_FAMILY, size=11, bold=True, color="1B2234")

        # Fills
        fill_header = PatternFill(start_color="1B2234", end_color="1B2234", fill_type="solid")
        fill_alt_row = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_highlight_total = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")

        # Borders
        thin_border_side = Side(border_style="thin", color="CBD5E1")
        thin_border = Border(
            left=thin_border_side, right=thin_border_side,
            top=thin_border_side, bottom=thin_border_side
        )
        double_bottom_border = Border(
            left=thin_border_side, right=thin_border_side,
            top=thin_border_side, bottom=Side(border_style="double", color="FF6B35")
        )

        # 1. Company Header
        ws.merge_cells("A1:H1")
        ws["A1"] = f"{settings.COMPANY_NAME} & PCCC"
        ws["A1"].font = font_header_company
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells("A2:H2")
        ws["A2"] = f"VP & Nhà Máy: {settings.COMPANY_ADDRESS} | Hotline PCCC: {settings.COMPANY_HOTLINE} | Email: {settings.COMPANY_EMAIL}"
        ws["A2"].font = font_sub_company
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells("A3:H3")
        ws["A3"] = f"Website: {settings.COMPANY_WEBSITE} | Lĩnh vực: Thiết Bị PCCC, Van Ngăn Cháy, Ống Gió & Cơ Điện Công Trình"
        ws["A3"].font = font_sub_company
        ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

        # 2. Quotation Title (Multi-language)
        ws.merge_cells("A5:H5")
        ws["A5"] = t("excel_title", lang)
        ws["A5"].font = font_title
        ws["A5"].alignment = Alignment(horizontal="center", vertical="center")

        # 3. Customer & Quote Info
        ws["A7"] = f"{t('excel_customer', lang)}:"
        ws["A7"].font = font_bold
        ws["B7"] = quote.customer_name
        ws["B7"].font = font_bold

        ws["A8"] = f"{t('excel_project', lang)}:"
        ws["A8"].font = font_regular
        ws["B8"] = quote.project_name
        ws["B8"].font = font_regular

        ws["A9"] = f"{t('customer_phone', lang)}:"
        ws["A9"].font = font_regular
        ws["B9"] = quote.customer_phone or "---"
        ws["B9"].font = font_regular

        ws["A10"] = f"{t('project_address', lang)}:"
        ws["A10"].font = font_regular
        ws["B10"] = quote.project_address or "---"
        ws["B10"].font = font_regular

        ws["F7"] = f"{t('excel_quote_no', lang)}:"
        ws["F7"].font = font_bold
        ws["G7"] = quote.quote_code
        ws["G7"].font = font_bold

        ws["F8"] = f"{t('excel_date', lang)}:"
        ws["F8"].font = font_regular
        ws["G8"] = quote.created_at[:10]
        ws["G8"].font = font_regular

        ws["F9"] = f"Validity:" if lang != "vi" else "Hiệu lực:"
        ws["F9"].font = font_regular
        ws["G9"] = f"{settings.QUOTE_VALIDITY_DAYS} days / ngày"
        ws["G9"].font = font_regular

        ws["F10"] = "Phụ trách / Contact:"
        ws["F10"].font = font_regular
        ws["G10"] = "Vertex Sales & Engineering"
        ws["G10"].font = font_regular

        # 4. Table Header (Multi-language)
        brand_col_title = "Hãng SX / Xuất Xứ" if lang == "vi" else "Brand / Origin"
        headers = [
            ("A12", t("col_stt", lang), 6),
            ("B12", t("col_item_code", lang), 14),
            ("C12", t("col_item_name", lang), 38),
            ("D12", brand_col_title, 20),
            ("E12", t("col_unit", lang), 8),
            ("F12", t("col_quantity", lang), 14),
            ("G12", f"{t('col_unit_price', lang)} (VNĐ)", 16),
            ("H12", f"{t('col_total_price', lang)} (VNĐ)", 18),
            ("I12", t("col_notes", lang), 24)
        ]

        ws.row_dimensions[12].height = 26
        for cell_ref, text, col_width in headers:
            cell = ws[cell_ref]
            cell.value = text
            cell.font = font_table_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # 5. Table Data Rows
        current_row = 13
        for idx, item in enumerate(quote.items, start=1):
            ws.row_dimensions[current_row].height = 22
            row_fill = fill_alt_row if idx % 2 == 0 else PatternFill(fill_type=None)

            ws[f"A{current_row}"] = item.stt
            ws[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"B{current_row}"] = item.item_code
            ws[f"B{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            full_name = item.item_name
            if item.spec and item.spec not in full_name:
                full_name = f"{item.item_name} ({item.spec})"
            ws[f"C{current_row}"] = full_name
            ws[f"C{current_row}"].alignment = Alignment(horizontal="left", vertical="center")

            ws[f"D{current_row}"] = item.brand or "Vertex Standard"
            ws[f"D{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"E{current_row}"] = item.unit
            ws[f"E{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            qty_val = item.area_m2 if item.area_m2 > 0 else item.quantity
            ws[f"F{current_row}"] = qty_val
            ws[f"F{current_row}"].number_format = "#,##0.00" if item.area_m2 > 0 else "#,##0"
            ws[f"F{current_row}"].alignment = Alignment(horizontal="right", vertical="center")

            ws[f"G{current_row}"] = item.unit_price
            ws[f"G{current_row}"].number_format = "#,##0"
            ws[f"G{current_row}"].alignment = Alignment(horizontal="right", vertical="center")

            ws[f"H{current_row}"] = item.total_price
            ws[f"H{current_row}"].number_format = "#,##0"
            ws[f"H{current_row}"].alignment = Alignment(horizontal="right", vertical="center")

            note_str = item.notes
            if item.labor_description and item.labor_unit_cost > 0:
                note_str = f"{item.labor_description} | {item.notes}".strip(" |")
            ws[f"I{current_row}"] = note_str
            ws[f"I{current_row}"].alignment = Alignment(horizontal="left", vertical="center")

            for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
                c = ws[f"{col_letter}{current_row}"]
                c.font = font_regular
                c.border = thin_border
                if row_fill.fill_type:
                    c.fill = row_fill

            current_row += 1

        # 6. Financial Summary Block (Multi-language)
        # Subtotal
        ws.merge_cells(f"A{current_row}:G{current_row}")
        ws[f"A{current_row}"] = f"{t('excel_subtotal', lang)}:"
        ws[f"A{current_row}"].font = font_bold
        ws[f"A{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"H{current_row}"] = quote.subtotal
        ws[f"H{current_row}"].font = font_bold
        ws[f"H{current_row}"].number_format = "#,##0"
        ws[f"H{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
        for col_l in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws[f"{col_l}{current_row}"].border = thin_border
        current_row += 1

        # Discount
        ws.merge_cells(f"A{current_row}:G{current_row}")
        disc_percent = int(quote.discount_rate * 100)
        ws[f"A{current_row}"] = f"{t('excel_discount', lang)} ({disc_percent}%):"
        ws[f"A{current_row}"].font = font_regular
        ws[f"A{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"H{current_row}"] = -quote.discount_amount
        ws[f"H{current_row}"].font = font_regular
        ws[f"H{current_row}"].number_format = "#,##0"
        ws[f"H{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
        for col_l in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws[f"{col_l}{current_row}"].border = thin_border
        current_row += 1

        # VAT
        ws.merge_cells(f"A{current_row}:G{current_row}")
        vat_percent = int(quote.vat_rate * 100)
        ws[f"A{current_row}"] = f"{t('excel_vat', lang)} ({vat_percent}%):"
        ws[f"A{current_row}"].font = font_regular
        ws[f"A{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"H{current_row}"] = quote.vat_amount
        ws[f"H{current_row}"].font = font_regular
        ws[f"H{current_row}"].number_format = "#,##0"
        ws[f"H{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
        for col_l in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws[f"{col_l}{current_row}"].border = thin_border
        current_row += 1

        # Grand Total
        ws.merge_cells(f"A{current_row}:G{current_row}")
        ws[f"A{current_row}"] = f"{t('excel_grand_total', lang)}:"
        ws[f"A{current_row}"].font = font_total_bold
        ws[f"A{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"H{current_row}"] = quote.total_amount
        ws[f"H{current_row}"].font = font_total_bold
        ws[f"H{current_row}"].number_format = "#,##0"
        ws[f"H{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"H{current_row}"].fill = fill_highlight_total
        ws.row_dimensions[current_row].height = 25

        for col_l in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws[f"{col_l}{current_row}"].border = double_bottom_border
        current_row += 1

        # In words
        ws.merge_cells(f"A{current_row}:I{current_row}")
        ws[f"A{current_row}"] = f"{t('excel_amount_in_words', lang)}: {quote.total_amount_in_words}"
        ws[f"A{current_row}"].font = Font(name=FONT_FAMILY, size=10, italic=True, bold=True, color="FF6B35")
        ws[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="center")
        current_row += 2

        # 7. Commercial Terms
        ws.merge_cells(f"A{current_row}:I{current_row}")
        ws[f"A{current_row}"] = "TERMS & CONDITIONS / ĐIỀU KHOẢN THƯƠNG MẠI & BẢO HÀNH PCCC:"
        ws[f"A{current_row}"].font = font_bold
        current_row += 1

        terms = [
            f"1. Giá trên đã bao gồm thuế GTGT ({vat_percent}% VAT) và chi phí đóng gói kiểm định tiêu chuẩn Vertex PCCC.",
            "2. Thiết bị PCCC có đầy đủ chứng chỉ xuất xưởng (CO/CQ) và tem kiểm định PCCC của Cục Cảnh sát PCCC & CNCH.",
            "3. Thời gian giao hàng: 03 - 05 ngày làm việc kể từ ngày xác nhận đơn hàng.",
            "4. Điều kiện thanh toán: Tạm ứng 30% khi đặt hàng, thanh toán 70% trước khi nhận hàng/bàn giao.",
            f"5. Báo giá có hiệu lực trong vòng {settings.QUOTE_VALIDITY_DAYS} ngày kể từ ngày phát hành."
        ]

        for term in terms:
            ws.merge_cells(f"A{current_row}:H{current_row}")
            ws[f"A{current_row}"] = term
            ws[f"A{current_row}"].font = font_italic
            current_row += 1

        current_row += 1

        # 8. Signatures Block
        ws.merge_cells(f"A{current_row}:C{current_row}")
        ws[f"A{current_row}"] = t("excel_customer_sign", lang)
        ws[f"A{current_row}"].font = font_bold
        ws[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"E{current_row}:H{current_row}")
        ws[f"E{current_row}"] = f"VERTEX CONSTRUCTION & PCCC ({t('excel_approved_by', lang)})"
        ws[f"E{current_row}"].font = font_bold
        ws[f"E{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        for cell_ref, text, col_width in headers:
            col_letter = cell_ref[0]
            ws.column_dimensions[col_letter].width = col_width

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        wb.close()
        return output_path
