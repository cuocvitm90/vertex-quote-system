"""
Field Reports & GPS Attendance Management Router for Vertex Construction & PCCC
Provides Geolocation check-in tracking, Site Daily Reports, Supervisor Review,
and Excel Export for Managers and Field Engineers.
"""
import json
import uuid
import time
import io
import math
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Any
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, status, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import settings
from app.database.db import db
from app.database.models import (
    User, UserRole, AttendanceCheckin, AttendanceCheckinCreateRequest,
    FieldDailyReport, FieldDailyReportCreateRequest, FieldReportCommentRequest,
    GeofenceCheckRequest, GeofenceAlertRecord, GeofenceConfigUpdateRequest
)
from app.services.auth import get_current_user, get_current_user_optional, require_manager_or_admin
from app.services.file_validator import FileValidator

router = APIRouter(tags=["Field Reports & Attendance"])
templates = Jinja2Templates(directory="app/templates")

# Vertex Project Construction Sites Reference with Configurable Geofence Radius (meters)
PROJECT_SITES = [
    {
        "id": "site_delta_grand",
        "name": "Khách Sạn 5 Sao Delta Grand (Hà Nội)",
        "address": "Lô B2, KĐT Ngoại Giao Đoàn, Bắc Từ Liêm, Hà Nội",
        "lat": 21.0568,
        "lng": 105.7925,
        "radius_meters": 200.0
    },
    {
        "id": "site_masterise_marina",
        "name": "Khu Căn Hộ Masterise Marina (TP.HCM)",
        "address": "Số 2 Tôn Đức Thắng, P. Bến Nghé, Quận 1, TP.HCM",
        "lat": 10.7826,
        "lng": 106.7029,
        "radius_meters": 200.0
    },
    {
        "id": "site_dhg_pharma",
        "name": "Nhà Máy Dược Phẩm Dược Hậu Giang (Bình Dương)",
        "address": "Đường số 6, KCN VSIP 1, Thuận An, Bình Dương",
        "lat": 10.9582,
        "lng": 106.6985,
        "radius_meters": 250.0
    },
    {
        "id": "site_vertex_tower",
        "name": "Văn Phòng Trụ Sở Vertex Construction & PCCC",
        "address": "Tầng 5, Tòa Nhà Vertex Tower, Cầu Giấy, Hà Nội",
        "lat": 21.0285,
        "lng": 105.7823,
        "radius_meters": 150.0
    }
]


def calculate_haversine_distance_meters(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """
    Calculates the great circle distance between two points on the earth in meters
    using the Haversine spherical formula.
    """
    R = 6371000.0  # Earth radius in meters
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)

    a = (math.sin(delta_phi / 2.0) ** 2) + \
        (math.cos(phi1) * math.cos(phi2) * (math.sin(delta_lambda / 2.0) ** 2))
    c = 2.0 * math.atan2(math.sqrt(a), math.sqrt(1.0 - a))

    return round(R * c, 1)


def evaluate_geofence_status(
    user_lat: float,
    user_lng: float,
    site_dict: Dict[str, Any],
    checkin_type: str = "IN",
    notes: Optional[str] = ""
) -> Tuple[str, float, bool, str]:
    """
    Evaluates whether a GPS position is inside or outside site geofence radius.
    Returns: (status, distance_meters, is_alert, alert_message)
    """
    site_lat = float(site_dict["lat"])
    site_lng = float(site_dict["lng"])
    radius = float(site_dict.get("radius_meters", 200.0))

    dist_meters = calculate_haversine_distance_meters(user_lat, user_lng, site_lat, site_lng)

    # 1. Within safe geofence radius
    if dist_meters <= radius:
        return "ON_SITE", dist_meters, False, f"Vị trí hợp lệ, nằm trong bán kính an toàn {radius:.0f}m của dự án (cách tâm {dist_meters:.0f}m)."

    # 2. Outside radius: Check for business survey or procurement exemptions
    clean_notes = (notes or "").lower()
    is_survey = checkin_type == "SITE_VISIT" or any(
        kw in clean_notes for kw in ["khảo sát", "vật tư", "mua vật tư", "nghiệm thu", "công tác", "họp", "thẩm duyệt"]
    )

    if is_survey:
        return "SITE_VISIT_APPROVED", dist_meters, False, f"Check-in ngoài vùng có lý do khảo sát/công tác hợp lệ (cách tâm {dist_meters:.0f}m)."

    # 3. Geofence breach -> Trigger Out-of-Zone Alert
    msg = (
        f"⚠️ CẢNH BÁO RANH GIỚI: Nhân sự check-in ngoài phạm vi an toàn công trình '{site_dict['name']}'! "
        f"Vị trí lệch: {dist_meters:.0f}m (vượt quá bán kính cho phép {radius:.0f}m)."
    )
    return "OUT_OF_ZONE", dist_meters, True, msg


@router.get("/field-reports", response_class=HTMLResponse)
async def serve_field_reports_page(
    request: Request,
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """
    Renders the dedicated Field Reports, GPS Attendance & Geofencing Live Tracking Workspace.
    Enforces authentication: Redirects to /login if unauthenticated.
    """
    if not current_user:
        return RedirectResponse(url="/login", status_code=status.HTTP_302_FOUND)

    all_users = db.list_all_users()
    approved_users = [u for u in all_users if u.status.value == "ACTIVE" and u.is_active]
    checkins = db.list_checkins(limit=50)
    reports = db.list_field_reports(limit=50)
    geofence_alerts = db.list_geofence_alerts(limit=20)
    
    return templates.TemplateResponse(
        request=request,
        name="field_reports.html",
        context={
            "request": request,
            "settings": settings,
            "user": current_user,
            "approved_users": approved_users,
            "checkins": checkins,
            "reports": reports,
            "geofence_alerts": geofence_alerts,
            "project_sites": PROJECT_SITES
        }
    )


@router.get("/api/field/personnel")
async def get_field_personnel(current_user: User = Depends(get_current_user)):
    """
    Returns dynamically synchronized list of approved personnel from User Management
    """
    all_users = db.list_all_users()
    approved_users = [u for u in all_users if u.status.value == "ACTIVE" and u.is_active]
    
    role_labels = {
        "ADMIN": "Giám Đốc / Ban Lãnh Đạo",
        "MANAGER": "Trưởng Phòng Kinh Doanh / QS",
        "STAFF": "Kỹ Sư QS / Giám Sát Hiện Trường",
        "DEALER": "Đại Lý Phân Phối",
        "PARTNER": "Đối Tác Nhà Thầu / CĐT"
    }
    
    data = []
    for u in approved_users:
        data.append({
            "id": u.id,
            "username": u.username,
            "full_name": u.full_name,
            "role": u.role.value,
            "role_label": role_labels.get(u.role.value, u.role.value),
            "company_name": u.company_name or "Công Ty Cổ Phần PCCC Vertex",
            "phone": u.phone or "Chưa cập nhật",
            "email": u.email or "",
            "status": u.status.value,
            "created_at": u.created_at
        })
        
    return JSONResponse(content={
        "status": "success",
        "total_personnel": len(data),
        "data": data
    })


@router.get("/api/field/sites")
async def get_project_sites(current_user: User = Depends(get_current_user)):
    """Returns list of active Vertex construction project sites for check-in & geofencing"""
    return JSONResponse(content={"status": "success", "sites": PROJECT_SITES})


@router.post("/api/field/geofence-check")
async def check_geofence_boundary(
    req: GeofenceCheckRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Real-time Geofence Boundary Check API.
    Calculates distance in meters, tests against radius boundary, and returns status.
    """
    site = next((s for s in PROJECT_SITES if s["name"] == req.project_site or s["id"] == req.project_site), PROJECT_SITES[0])
    status_label, distance_m, is_alert, msg = evaluate_geofence_status(
        user_lat=req.latitude,
        user_lng=req.longitude,
        site_dict=site,
        checkin_type=req.checkin_type or "IN",
        notes=req.notes or ""
    )

    return JSONResponse(content={
        "status": "success",
        "geofence_status": status_label,
        "distance_meters": distance_m,
        "radius_meters": site.get("radius_meters", 200.0),
        "is_alert": is_alert,
        "message": msg,
        "site_center": {"lat": site["lat"], "lng": site["lng"], "name": site["name"]}
    })


@router.get("/api/field/geofence-alerts")
async def list_geofence_alerts(
    current_user: User = Depends(get_current_user)
):
    """Lists out-of-zone geofencing alerts for managerial audit"""
    alerts = db.list_geofence_alerts(limit=50)
    return JSONResponse(content={
        "status": "success",
        "total_alerts": len(alerts),
        "data": [a.model_dump() for a in alerts]
    })


@router.post("/api/field/geofence-config")
async def update_geofence_config(
    req: GeofenceConfigUpdateRequest,
    current_user: User = Depends(require_manager_or_admin)
):
    """Allows Managers/Admins to configure safe geofence radius and GPS coordinates for a project site"""
    for s in PROJECT_SITES:
        if s["name"] == req.project_site or s["id"] == req.project_site:
            if req.radius_meters is not None:
                s["radius_meters"] = max(50.0, min(5000.0, float(req.radius_meters)))
            if req.lat is not None:
                s["lat"] = float(req.lat)
            if req.lng is not None:
                s["lng"] = float(req.lng)
            if req.address:
                s["address"] = req.address
            return JSONResponse(content={
                "status": "success",
                "message": f"Đã cập nhật cấu hình tọa độ & bán kính an toàn ({s['radius_meters']:.0f}m) của '{s['name']}'!",
                "site": s
            })
    
    # If new site, create it
    if req.lat is not None and req.lng is not None:
        new_site = {
            "id": f"site_{uuid.uuid4().hex[:8]}",
            "name": req.project_site,
            "address": req.address or f"Tọa độ: {req.lat:.4f}, {req.lng:.4f}",
            "lat": float(req.lat),
            "lng": float(req.lng),
            "radius_meters": float(req.radius_meters) if req.radius_meters else 200.0
        }
        PROJECT_SITES.append(new_site)
        return JSONResponse(content={
            "status": "success",
            "message": f"Đã thêm mới tọa độ công trình '{new_site['name']}' thành công!",
            "site": new_site
        })

    raise HTTPException(status_code=404, detail="Không tìm thấy dự án này!")



@router.post("/api/field/geofence-alerts/{alert_id}/resolve")
async def resolve_geofence_alert(
    alert_id: str,
    current_user: User = Depends(require_manager_or_admin)
):
    """Manager resolves/dismisses a geofencing alert incident"""
    success = db.resolve_geofence_alert(alert_id, status="RESOLVED")
    if not success:
        raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi cảnh báo này!")
    return JSONResponse(content={"status": "success", "message": "Đã xử lý cảnh báo ranh giới thành công!"})


@router.get("/api/field/checkins")
async def get_checkins(
    user_only: bool = False,
    current_user: User = Depends(get_current_user)
):
    """Lists GPS check-in history"""
    uid = current_user.id if (user_only and current_user.role not in [UserRole.ADMIN, UserRole.MANAGER]) else None
    checkins = db.list_checkins(limit=100, user_id=uid)
    return JSONResponse(content={"status": "success", "data": [c.model_dump() for c in checkins]})


@router.post("/api/field/checkin")
async def submit_gps_checkin(
    req: AttendanceCheckinCreateRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Submits a GPS Geolocation Check-in with automated Geofencing verification.
    Calculates precise distance to project site and tags status (ON_SITE, OUT_OF_ZONE, SITE_VISIT_APPROVED).
    Triggers automated Manager & Admin alert if geofence is breached.
    """
    checkin_id = f"chk-{int(time.time())}-{uuid.uuid4().hex[:6]}"
    
    # 1. Match project site
    site = next((s for s in PROJECT_SITES if s["name"] == req.project_site), PROJECT_SITES[0])
    address = req.address_resolved or site["address"]

    # 2. Evaluate Geofence boundary
    geofence_status, distance_m, is_alert, alert_msg = evaluate_geofence_status(
        user_lat=req.latitude,
        user_lng=req.longitude,
        site_dict=site,
        checkin_type=req.checkin_type,
        notes=req.notes or ""
    )

    # 3. Create Attendance checkin record
    checkin_item = db.create_checkin({
        "id": checkin_id,
        "user_id": current_user.id,
        "user_name": current_user.full_name,
        "project_site": req.project_site,
        "checkin_type": req.checkin_type,
        "latitude": req.latitude,
        "longitude": req.longitude,
        "accuracy_meters": req.accuracy_meters or 10.0,
        "address_resolved": address,
        "status": geofence_status,
        "notes": req.notes or "",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })

    # 4. If Geofence Breached -> Record Alert and Trigger Notification Log
    alert_record = None
    if is_alert:
        alert_id = f"gfa-{int(time.time())}-{uuid.uuid4().hex[:6]}"
        alert_record = db.create_geofence_alert({
            "id": alert_id,
            "user_id": current_user.id,
            "user_name": current_user.full_name,
            "project_site": req.project_site,
            "latitude": req.latitude,
            "longitude": req.longitude,
            "distance_meters": distance_m,
            "radius_meters": site.get("radius_meters", 200.0),
            "alert_message": alert_msg,
            "status": "UNRESOLVED",
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        # Record into Immutable Audit Trail
        db.add_audit_log(
            quote_id="",
            user_id=current_user.id,
            user_name=current_user.full_name,
            user_role=current_user.role.value,
            action="GEOFENCE_BREACH_ALERT",
            details=f"Nhân viên {current_user.full_name} check-in ngoài ranh giới {req.project_site} (cách {distance_m:.0f}m, bán kính {site.get('radius_meters', 200):.0f}m)."
        )

    return JSONResponse(content={
        "status": "success",
        "geofence_status": geofence_status,
        "distance_meters": distance_m,
        "radius_meters": site.get("radius_meters", 200.0),
        "is_alert": is_alert,
        "alert_message": alert_msg,
        "message": f"Check-in thành công ({geofence_status}) tại {req.project_site} lúc {datetime.now().strftime('%H:%M:%S')}!",
        "data": checkin_item.model_dump()
    })



@router.get("/api/field/reports")
async def get_field_reports(
    project_name: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Lists field daily reports"""
    reports = db.list_field_reports(limit=100, project_name=project_name)
    return JSONResponse(content={"status": "success", "data": [r.model_dump() for r in reports]})


@router.post("/api/field/report")
async def submit_field_daily_report(
    req: FieldDailyReportCreateRequest,
    current_user: User = Depends(get_current_user)
):
    """Submits a new daily construction / installation report"""
    from app.services.sanitizer import clean_string

    report_id = f"rep-{int(time.time())}-{uuid.uuid4().hex[:6]}"
    report_date = req.report_date or datetime.now().strftime("%Y-%m-%d")

    clean_project = clean_string(req.project_name, escape_html_entities=False)
    clean_summary = clean_string(req.work_summary, escape_html_entities=False)
    clean_issues = clean_string(req.issues_and_risks or "", escape_html_entities=False)
    clean_next = clean_string(req.next_plan or "", escape_html_entities=False)
    clean_weather = clean_string(req.weather_condition or "Nắng ráo", escape_html_entities=False)

    report_item = db.create_field_report({
        "id": report_id,
        "user_id": current_user.id,
        "user_name": current_user.full_name,
        "project_name": clean_project,
        "report_date": report_date,
        "weather_condition": clean_weather,
        "work_summary": clean_summary,
        "progress_percent": req.progress_percent,
        "workforce_count": req.workforce_count,
        "issues_and_risks": clean_issues,
        "next_plan": clean_next,
        "photos_json": req.photos_json or "[]",
        "supervisor_comment": "",
        "status": "SUBMITTED",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })

    return JSONResponse(content={
        "status": "success",
        "message": f"Đã gửi báo cáo công việc dự án {clean_project} thành công!",
        "data": report_item.model_dump()
    })


@router.post("/api/field/report/{report_id}/comment")
async def comment_on_field_report(
    report_id: str,
    req: FieldReportCommentRequest,
    current_user: User = Depends(require_manager_or_admin)
):
    """Manager/Admin adds review comments and approves field report"""
    from app.services.sanitizer import clean_string
    clean_comment = clean_string(req.comment, escape_html_entities=False)

    updated = db.update_field_report_comment(
        report_id=report_id,
        comment=f"[{current_user.full_name}] {clean_comment}",
        status=req.status or "APPROVED"
    )
    if not updated:
        raise HTTPException(status_code=404, detail="Không tìm thấy báo cáo yêu cầu.")

    return JSONResponse(content={
        "status": "success",
        "message": "Đã lưu nhận xét và phê duyệt báo cáo hiện trường!",
        "data": updated.model_dump()
    })


@router.get("/api/field/export-excel")
async def export_field_reports_excel(
    current_user: User = Depends(get_current_user)
):
    """Exports all attendance records and field daily reports into a branded Excel spreadsheet"""
    wb = Workbook()
    
    # Sheet 1: Check-in Attendance
    ws1 = wb.active
    ws1.title = "Nhat_Ky_Cham_Cong_GPS"
    ws1.views.sheetView[0].showGridLines = True

    # Styling tokens
    navy_fill = PatternFill(start_color="1B2234", end_color="1B2234", fill_type="solid")
    orange_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    gray_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=15, bold=True, color="1B2234")
    header_font = Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
    regular_font = Font(name="Calibri", size=10, color="0F172A")
    bold_font = Font(name="Calibri", size=10, bold=True, color="0F172A")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    # Title Banner
    ws1.merge_cells("A1:G1")
    t1 = ws1.cell(row=1, column=1, value="BẢNG TỔNG HỢP CHẤM CÔNG GPS HIỆN TRƯỜNG - VERTEX PCCC")
    t1.font = title_font
    t1.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    headers1 = ["STT", "Thời Gian", "Nhân Viên", "Dự Án / Công Trình", "Loại", "Tọa Độ GPS", "Ghi Chú"]
    ws1.row_dimensions[3].height = 22
    for col_idx, h in enumerate(headers1, start=1):
        c = ws1.cell(row=3, column=col_idx, value=h)
        c.fill = navy_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    checkins = db.list_checkins(limit=500)
    for idx, item in enumerate(checkins, start=1):
        r = idx + 3
        ws1.row_dimensions[r].height = 19
        c1 = ws1.cell(row=r, column=1, value=idx)
        c2 = ws1.cell(row=r, column=2, value=item.created_at)
        c3 = ws1.cell(row=r, column=3, value=item.user_name)
        c4 = ws1.cell(row=r, column=4, value=item.project_site)
        c5 = ws1.cell(row=r, column=5, value="Vào ca" if item.checkin_type == "IN" else ("Tan ca" if item.checkin_type == "OUT" else "Khảo sát"))
        c6 = ws1.cell(row=r, column=6, value=f"{item.latitude:.4f}, {item.longitude:.4f}")
        c7 = ws1.cell(row=r, column=7, value=item.notes)

        for c in [c1, c2, c4, c5, c6, c7]:
            c.font = regular_font
        c3.font = bold_font

        for c in [c1, c2, c3, c4, c5, c6, c7]:
            c.border = thin_border
            if idx % 2 == 0:
                c.fill = gray_fill
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c5.alignment = Alignment(horizontal="center", vertical="center")
        c6.alignment = Alignment(horizontal="center", vertical="center")

    col_widths1 = {1: 8, 2: 18, 3: 25, 4: 35, 5: 12, 6: 22, 7: 35}
    for col_idx, width in col_widths1.items():
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # Sheet 2: Daily Reports
    ws2 = wb.create_sheet(title="Bao_Cao_Tien_Do_Hang_Ngay")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:H1")
    t2 = ws2.cell(row=1, column=1, value="BÁO CÁO CÔNG VIỆC HIỆN TRƯỜNG & TIẾN ĐỘ THI CÔNG PCCC")
    t2.font = title_font
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    headers2 = ["STT", "Ngày Báo Cáo", "Kỹ Sư Lập", "Dự Án", "Nội Dung Công Việc", "Tiến Độ (%)", "Quân Số", "Ý Kiến Quản Lý"]
    ws2.row_dimensions[3].height = 22
    for col_idx, h in enumerate(headers2, start=1):
        c = ws2.cell(row=3, column=col_idx, value=h)
        c.fill = orange_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    reports = db.list_field_reports(limit=500)
    for idx, rep in enumerate(reports, start=1):
        r = idx + 3
        ws2.row_dimensions[r].height = 22
        c1 = ws2.cell(row=r, column=1, value=idx)
        c2 = ws2.cell(row=r, column=2, value=rep.report_date)
        c3 = ws2.cell(row=r, column=3, value=rep.user_name)
        c4 = ws2.cell(row=r, column=4, value=rep.project_name)
        c5 = ws2.cell(row=r, column=5, value=rep.work_summary)
        c6 = ws2.cell(row=r, column=6, value=f"{rep.progress_percent}%")
        c7 = ws2.cell(row=r, column=7, value=rep.workforce_count)
        c8 = ws2.cell(row=r, column=8, value=rep.supervisor_comment or "Chưa có nhận xét")

        for c in [c1, c2, c4, c5, c6, c7, c8]:
            c.font = regular_font
        c3.font = bold_font

        for c in [c1, c2, c3, c4, c5, c6, c7, c8]:
            c.border = thin_border
            if idx % 2 == 0:
                c.fill = gray_fill
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c6.alignment = Alignment(horizontal="right", vertical="center")
        c7.alignment = Alignment(horizontal="center", vertical="center")

    col_widths2 = {1: 8, 2: 15, 3: 25, 4: 30, 5: 45, 6: 14, 7: 12, 8: 35}
    for col_idx, width in col_widths2.items():
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    # Sheet 3: Approved Personnel Roster (Synchronized from User Management)
    ws3 = wb.create_sheet(title="Danh_Sach_Nhan_Su")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells("A1:G1")
    t3 = ws3.cell(row=1, column=1, value="DANH SÁCH CÁN BỘ & NHÂN SỰ ĐƯỢC PHÊ DUYỆT - VERTEX PCCC")
    t3.font = title_font
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    headers3 = ["STT", "Họ và Tên", "Tên Đăng Nhập", "Chức Vụ / Quyền Hạn", "Đơn Vị / Công Ty", "Số Điện Thoại", "Trạng Thái"]
    ws3.row_dimensions[3].height = 22
    for col_idx, h in enumerate(headers3, start=1):
        c = ws3.cell(row=3, column=col_idx, value=h)
        c.fill = navy_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    role_labels = {
        "ADMIN": "Giám Đốc / Ban Lãnh Đạo",
        "MANAGER": "Trưởng Phòng Kinh Doanh / QS",
        "STAFF": "Kỹ Sư QS / Giám Sát Hiện Trường",
        "DEALER": "Đại Lý Phân Phối",
        "PARTNER": "Đối Tác Nhà Thầu / CĐT"
    }

    all_users = db.list_all_users()
    approved_users = [u for u in all_users if u.status.value == "ACTIVE" and u.is_active]
    for idx, u in enumerate(approved_users, start=1):
        r = idx + 3
        ws3.row_dimensions[r].height = 20
        c1 = ws3.cell(row=r, column=1, value=idx)
        c2 = ws3.cell(row=r, column=2, value=u.full_name)
        c3 = ws3.cell(row=r, column=3, value=u.username)
        c4 = ws3.cell(row=r, column=4, value=role_labels.get(u.role.value, u.role.value))
        c5 = ws3.cell(row=r, column=5, value=u.company_name or "Công Ty Cổ Phần PCCC Vertex")
        c6 = ws3.cell(row=r, column=6, value=u.phone or "Chưa cập nhật")
        c7 = ws3.cell(row=r, column=7, value="🟢 ĐÃ PHÊ DUYỆT (ACTIVE)")

        for c in [c1, c3, c4, c5, c6, c7]:
            c.font = regular_font
        c2.font = bold_font

        for c in [c1, c2, c3, c4, c5, c6, c7]:
            c.border = thin_border
            if idx % 2 == 0:
                c.fill = gray_fill
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c3.alignment = Alignment(horizontal="center", vertical="center")
        c6.alignment = Alignment(horizontal="center", vertical="center")
        c7.alignment = Alignment(horizontal="center", vertical="center")

    col_widths3 = {1: 8, 2: 28, 3: 20, 4: 32, 5: 35, 6: 18, 7: 26}
    for col_idx, width in col_widths3.items():
        ws3.column_dimensions[get_column_letter(col_idx)].width = width

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"Vertex_Bao_Cao_Cham_Cong_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
